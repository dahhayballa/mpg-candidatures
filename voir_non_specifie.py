#!/usr/bin/env python3
"""
MPG — Export des candidats "Non spécifié"
Pour vérification manuelle par la commission
"""
import sqlite3, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

DB_PATH = "candidates.db"

def thin():
    s = Side(style='thin', color="BDC3C7")
    return Border(left=s, right=s, top=s, bottom=s)

def cs(cell, val, bg, fg="000000", bold=False, size=9, align="center"):
    cell.value = val
    cell.font  = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill  = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = thin()

conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row
c = conn.cursor()

cols = [r[1] for r in c.execute("PRAGMA table_info(candidates)").fetchall()]
has_hd    = "hors_delai"  in cols
has_score = "score_total" in cols
has_ia    = "ia_scored"   in cols
has_cm    = "contenu_manquant" in cols

where = "(specialty='Non spécifié' OR specialty IS NULL OR specialty='')"
if has_hd:
    where += " AND hors_delai=0"

extra = ""
if has_score: extra += ", score_total, mention"
if has_ia:    extra += ", ia_scored"
if has_cm:    extra += ", contenu_manquant"

rows = c.execute(f"""
    SELECT id, name, email_addr, specialty, status,
           has_cv, has_motivation, has_id, has_diplomas,
           num_emails, first_date, last_date,
           (has_cv + has_motivation + has_id + has_diplomas) as doc_count,
           body_preview
           {extra}
    FROM candidates
    WHERE {where}
    ORDER BY
        (has_cv + has_motivation + has_id + has_diplomas) DESC,
        first_date ASC
""").fetchall()
conn.close()

total = len(rows)
print(f"  Non spécifié : {total} candidats")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Non Spécifié"

# Titre
ws.merge_cells("A1:M1")
cs(ws["A1"],
   "EETFP-MPG — Candidats 'Non Spécifié' — Vérification manuelle de la commission",
   "1E3A5F", "FFFFFF", True, 12)
ws.row_dimensions[1].height = 26

ws.merge_cells("A2:M2")
cs(ws["A2"],
   f"{total} candidats — triés par qualité du dossier (complets en premier) puis date d'envoi",
   "EBF5FB", "1A5276", False, 9)
ws.row_dimensions[2].height = 14

ws.merge_cells("A3:M3")
cs(ws["A3"],
   "⚠  Ces candidats n'ont pas précisé leur filière. La commission peut leur attribuer une filière "
   "manuellement dans la colonne 'Filière suggérée' et les intégrer à la liste finale.",
   "FFFDE7", "7D6608", False, 9)
ws.row_dimensions[3].height = 24

# En-têtes
hdrs = ["N°", "Nom et Prénom", "E-mail", "Docs\n/4",
        "CV", "Lettre", "CIN", "Diplômes",
        "Statut", "Score\nIA", "Date dépôt",
        "Aperçu contenu", "Filière suggérée\n(à remplir)"]
for ci, h in enumerate(hdrs, 1):
    cs(ws.cell(row=5, column=ci), h, "1E3A5F", "FFFFFF", True, 9)
ws.row_dimensions[5].height = 30

# Données
for i, r in enumerate(rows):
    d      = dict(r)
    row_n  = 6 + i
    docs   = int(d.get("doc_count") or 0)
    score  = float(d.get("score_total") or 0) if has_score else 0
    cm     = bool(d.get("contenu_manquant") or 0) if has_cm else False
    status = d.get("status") or "—"

    # Couleur selon qualité dossier
    if docs == 4:   bg = "D5F5E3" if i%2==0 else "C8F7D6"  # vert
    elif docs >= 2: bg = "FEF9E7" if i%2==0 else "FDEBD0"  # jaune
    else:           bg = "FDEDEC" if i%2==0 else "FADBD8"  # rouge clair

    apercu = str(d.get("body_preview") or "")[:80]
    if cm:
        apercu = "⚠ Contenu illisible — " + apercu

    row_d = [
        i + 1,
        d.get("name") or "—",
        d.get("email_addr") or "—",
        f"{docs}/4",
        "✓" if d.get("has_cv")         else "✗",
        "✓" if d.get("has_motivation") else "✗",
        "✓" if d.get("has_id")         else "✗",
        "✓" if d.get("has_diplomas")   else "✗",
        status,
        f"{score:.0f}/100" if has_score and score > 0 else "—",
        d.get("first_date") or "—",
        apercu,
        "",  # Colonne vide pour filière suggérée
    ]
    for ci, val in enumerate(row_d, 1):
        cell = ws.cell(row=row_n, column=ci)
        al   = "left" if ci in (2, 3, 12, 13) else "center"
        fg   = "000000"
        if ci in (5,6,7,8):
            fg = "1E8449" if val == "✓" else "C0392B"
        elif ci == 4:
            fg = "1E8449" if docs==4 else ("D68910" if docs>=2 else "C0392B")
        cs(cell, val, bg, fg, bold=(ci in (1,2,4)), size=9, align=al)
    ws.row_dimensions[row_n].height = 15

# Résumé par qualité
sum_row = 6 + total + 2
complets  = sum(1 for r in rows if int(dict(r).get("doc_count") or 0) == 4)
partiels  = sum(1 for r in rows if int(dict(r).get("doc_count") or 0) in (2,3))
incomplets= sum(1 for r in rows if int(dict(r).get("doc_count") or 0) <= 1)

ws.merge_cells(f"A{sum_row}:M{sum_row}")
cs(ws[f"A{sum_row}"],
   f"RÉSUMÉ : {complets} dossiers complets (4/4) ✅  |  "
   f"{partiels} dossiers partiels (2-3/4) 🟡  |  "
   f"{incomplets} dossiers incomplets (0-1/4) ❌  |  "
   f"TOTAL : {total}",
   "1E3A5F", "FFFFFF", True, 10)
ws.row_dimensions[sum_row].height = 20

# Largeurs
for ci, w in enumerate([5,30,30,6,5,6,5,8,10,8,18,45,22], 1):
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.freeze_panes = "A6"

fname = f"Non_Specifie_MPG_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
wb.save(fname)

print(f"  ✅ Fichier généré : {fname}")
print(f"  Complets  (4/4) : {complets}")
print(f"  Partiels  (2-3) : {partiels}")
print(f"  Incomplets(0-1) : {incomplets}")
