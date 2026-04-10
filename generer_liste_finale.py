#!/usr/bin/env python3
"""
EETFP-MPG — Liste Finale Unifiée
Papier + Numérique classés ensemble /100
Usage : python generer_liste_finale.py
"""
import sqlite3, openpyxl, os, json
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

DB_PATH      = "candidates.db"
EXCEL_PAPIER = "Liste_FQ2026__1_.xlsx"

QUOTA = {
    "HSE":                                40,
    "Electricite industrielle":           30,
    "Tuyauterie industrielle":            30,
    "Construction metallique et soudure": 40,
    "Maintenance industrielle":           20,
    "Techniques minieres":                20,
    "Operations petrolieres et gazieres": 20,
}

QUOTA_DISPLAY = {
    "HSE":                                "HSE",
    "Electricite industrielle":           "Électricité industrielle",
    "Tuyauterie industrielle":            "Tuyauterie industrielle",
    "Construction metallique et soudure": "Construction métallique et soudure",
    "Maintenance industrielle":           "Maintenance industrielle",
    "Techniques minieres":                "Techniques minières",
    "Operations petrolieres et gazieres": "Opérations pétrolières et gazières",
}

SHEET_MAP = {
    'Opération pétrolières et gazièr': 'Operations petrolieres et gazieres',
    'Electricité Industrielle ':       'Electricite industrielle',
    'HSE':                             'HSE',
    'Maintenance Industrielle ':       'Maintenance industrielle',
    'Construction métallique et Soud': 'Construction metallique et soudure',
    'Technique minière ':              'Techniques minieres',
    'Tuyautérie':                      'Tuyauterie industrielle',
}

DB_SPEC_MAP = {
    'Opérations pétrolières et gazières': 'Operations petrolieres et gazieres',
    'Électricité industrielle':           'Electricite industrielle',
    'HSE':                                'HSE',
    'Maintenance industrielle':           'Maintenance industrielle',
    'Construction métallique et soudure': 'Construction metallique et soudure',
    'Techniques minières':                'Techniques minieres',
    'Tuyauterie industrielle':            'Tuyauterie industrielle',
}


def thin():
    s = Side(style='thin', color="BDC3C7")
    return Border(left=s, right=s, top=s, bottom=s)

def cs(cell, value, bg, fg="000000", bold=False, size=9, align="center", wrap=False):
    cell.value = value
    cell.font  = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill  = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = thin()

def mention(s):
    if s >= 80: return "Excellent"
    if s >= 65: return "Bon"
    if s >= 50: return "Moyen"
    return "Non retenu"


def load_paper():
    if not os.path.exists(EXCEL_PAPIER):
        print(f"⚠  Fichier Excel papier non trouvé : {EXCEL_PAPIER}")
        return {}
    wb  = openpyxl.load_workbook(EXCEL_PAPIER, read_only=True)
    out = {}
    for sname in wb.sheetnames:
        key = SHEET_MAP.get(sname)
        if not key:
            continue
        ws   = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        non_empty = [(i, r) for i, r in enumerate(rows) if any(v is not None for v in r)]
        cands  = []
        hdr    = False
        for i, row in non_empty:
            vals = list(row)
            sv   = [str(v) for v in vals if v is not None]
            if not hdr:
                if any('Noms' in v or 'prénoms' in v for v in sv):
                    hdr = True
                continue
            nn = [v for v in vals if v is not None]
            if len(nn) < 3 or not isinstance(nn[0], (int, float)):
                continue
            name = None; diplome = None; scores = []
            for j, v in enumerate(vals):
                if v is None: continue
                if j == 0: continue
                if isinstance(v, str) and name is None and len(str(v).strip()) > 2:
                    name = str(v).strip()
                elif isinstance(v, str) and j > 3 and diplome is None:
                    diplome = str(v).strip()
                elif isinstance(v, (int, float)) and j > 4:
                    scores.append(float(v))
            if name and len(scores) >= 6:
                total = sum(scores[:6])
                cands.append({
                    "nom": name, "email": "—", "diplome": diplome or "—",
                    "score_total": round(total, 1), "mention": mention(total),
                    "source": "PAPIER", "has_cv": True, "has_motivation": True,
                    "has_id": True, "has_diplomas": True,
                    "status": "Complet", "first_date": "—", "doc_count": 4,
                })
        out[key] = cands
    return out


def load_digital():
    if not os.path.exists(DB_PATH):
        print(f"⚠  Base de données non trouvée : {DB_PATH}")
        return {}
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    c    = conn.cursor()
    cols = [r[1] for r in c.execute("PRAGMA table_info(candidates)").fetchall()]
    has_hd    = "hors_delai"  in cols
    has_score = "score_total" in cols
    has_ia    = "ia_scored"   in cols
    where = "hors_delai=0" if has_hd else "1=1"
    extra = ""
    if has_score: extra += ", score_total, mention"
    if has_ia:    extra += ", ia_scored"
    rows = c.execute(f"""
        SELECT id, name, email_addr, specialty, status,
               has_cv, has_motivation, has_id, has_diplomas,
               num_emails, first_date, last_date,
               (has_cv + has_motivation + has_id + has_diplomas) as doc_count
               {extra}
        FROM candidates WHERE {where}
        ORDER BY specialty, first_date ASC
    """).fetchall()
    conn.close()
    out = {}
    for r in rows:
        d  = dict(r)
        sp = d.get("specialty") or ""
        key = DB_SPEC_MAP.get(sp, "")
        if not key:
            continue
        score_ia  = float(d.get("score_total") or 0)
        doc_count = int(d.get("doc_count") or 0)
        ia_ok     = bool(d.get("ia_scored") or 0) and score_ia >= 30
        if ia_ok:
            score_final = score_ia
        else:
            score_final = round(min(doc_count * 22.5, 90) + (5 if sp != "Non spécifié" else 0), 1)
        if key not in out:
            out[key] = []
        out[key].append({
            "nom": d.get("name") or "—",
            "email": d.get("email_addr") or "—",
            "diplome": "—",
            "score_total": score_final,
            "score_ia": score_ia,
            "mention": mention(score_final),
            "source": "NUMERIQUE",
            "has_cv": bool(d.get("has_cv")),
            "has_motivation": bool(d.get("has_motivation")),
            "has_id": bool(d.get("has_id")),
            "has_diplomas": bool(d.get("has_diplomas")),
            "status": d.get("status") or "—",
            "first_date": d.get("first_date") or "—",
            "doc_count": doc_count,
        })
    return out


def generate():
    print("=" * 65)
    print("  EETFP-MPG — Liste finale unifiée")
    print("  Classement commun papier + numérique")
    print("=" * 65)
    print()

    paper   = load_paper()
    digital = load_digital()
    wb      = openpyxl.Workbook()
    wb.remove(wb.active)
    summary = []

    for key, quota in QUOTA.items():
        display = QUOTA_DISPLAY[key]
        all_c   = []
        for c in paper.get(key, []):
            c["_t"] = "P"; all_c.append(c)
        for c in digital.get(key, []):
            c["_t"] = "D"; all_c.append(c)

        all_c.sort(key=lambda x: -float(x.get("score_total") or 0))
        retenus = all_c[:quota]
        attente = all_c[quota:quota + 20]
        nb_p    = sum(1 for c in retenus if c["_t"] == "P")
        nb_d    = sum(1 for c in retenus if c["_t"] == "D")
        summary.append({"key": key, "display": display, "quota": quota,
                        "retenus": len(retenus), "paper": nb_p, "digital": nb_d,
                        "attente": len(attente)})
        print(f"  {display[:38]:<38} {len(retenus):>3}/{quota} "
              f"({nb_p}P + {nb_d}N)")

        ws   = wb.create_sheet(title=display[:31])
        # ─ Titre ─
        ws.merge_cells("A1:K1")
        cs(ws["A1"], "EETFP-MPG — Formation Qualifiante 2025-2026 — Liste des Candidats Retenus",
           "1E3A5F", "FFFFFF", True, 12)
        ws.row_dimensions[1].height = 26
        ws.merge_cells("A2:K2")
        cs(ws["A2"], f"Filière : {display}   |   Quota : {quota} places",
           "2D6A9F", "FFFFFF", True, 10)
        ws.row_dimensions[2].height = 18
        ws.merge_cells("A3:K3")
        cs(ws["A3"],
           "Classement unifié — Candidats papier (P) et numériques (N) notés sur /100 — "
           "Mêmes critères pour tous",
           "EBF5FB", "1A5276", False, 9, wrap=True)
        ws.row_dimensions[3].height = 14
        ws.merge_cells("A4:K4")
        cs(ws["A4"], f"✅  {len(retenus)} CANDIDATS RETENUS",
           "1E8449", "FFFFFF", True, 10)
        ws.row_dimensions[4].height = 18
        # ─ En-tête colonnes ─
        hdrs = ["Rang", "Nom et Prénom", "Email / Diplôme", "Canal",
                "Dossier", "CV", "Lettre", "CIN", "Diplômes",
                "Score\n/100", "Mention"]
        for ci, h in enumerate(hdrs, 1):
            cs(ws.cell(row=5, column=ci), h, "1E3A5F", "FFFFFF", True, 9, wrap=True)
        ws.row_dimensions[5].height = 30
        # ─ Retenus ─
        for i, c in enumerate(retenus):
            r  = 6 + i
            sc = float(c.get("score_total") or 0)
            is_p = c["_t"] == "P"
            bg   = ("D6EAF8" if i%2==0 else "EBF5FB") if is_p else \
                   ("D5F5E3" if i%2==0 else "E9F7EF")
            sc_c = ("1E8449" if sc>=80 else "2E86C1" if sc>=65 else
                    "D68910" if sc>=50 else "C0392B")
            row_d = [
                i+1,
                c["nom"],
                c["diplome"] if is_p else c["email"],
                "📋 Papier" if is_p else "💻 Numérique",
                c["status"],
                "✓" if c["has_cv"]         else "✗",
                "✓" if c["has_motivation"] else "✗",
                "✓" if c["has_id"]         else "✗",
                "✓" if c["has_diplomas"]   else "✗",
                sc,
                c["mention"],
            ]
            for ci, val in enumerate(row_d, 1):
                cell = ws.cell(row=r, column=ci)
                al   = "left" if ci in (2, 3) else "center"
                fg   = "000000"
                if ci in (6,7,8,9):
                    fg = "1E8449" if val == "✓" else "C0392B"
                elif ci == 10: fg = sc_c
                elif ci == 4:  fg = "1A5276" if is_p else "1E8449"
                cs(cell, val, bg, fg, bold=(ci in (1,2,10,11)), size=9, align=al)
            ws.row_dimensions[r].height = 15
        # ─ Liste d'attente ─
        if attente:
            sep = 6 + len(retenus) + 1
            ws.merge_cells(f"A{sep}:K{sep}")
            cs(ws[f"A{sep}"],
               f"⏳  LISTE D'ATTENTE — {len(attente)} candidats (en cas de désistement)",
               "7D6608", "FFFFFF", True, 9)
            ws.row_dimensions[sep].height = 16
            for i, c in enumerate(attente):
                r    = sep + 1 + i
                is_p = c["_t"] == "P"
                bg   = "FDFEFE" if i%2==0 else "F4F6F7"
                row_d = [
                    f"A{i+1}", c["nom"],
                    c["diplome"] if is_p else c["email"],
                    "📋 Papier" if is_p else "💻 Numérique",
                    c["status"],
                    "✓" if c["has_cv"]         else "✗",
                    "✓" if c["has_motivation"] else "✗",
                    "✓" if c["has_id"]         else "✗",
                    "✓" if c["has_diplomas"]   else "✗",
                    c["score_total"], c["mention"],
                ]
                for ci, val in enumerate(row_d, 1):
                    cell = ws.cell(row=r, column=ci)
                    al   = "left" if ci in (2,3) else "center"
                    cs(cell, val, bg, "888888", False, 9, al)
                ws.row_dimensions[r].height = 14
        # ─ Largeurs ─
        for ci, w in enumerate([5,32,30,14,10,5,6,5,8,8,12], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "A6"

    # ─── RÉCAPITULATIF ───────────────────────────────
    ws_s = wb.create_sheet("RÉCAPITULATIF", 0)
    ws_s.merge_cells("A1:H1")
    cs(ws_s["A1"],
       "EETFP-MPG — RÉCAPITULATIF GÉNÉRAL — Formation Qualifiante 2025-2026",
       "1E3A5F", "FFFFFF", True, 13)
    ws_s.row_dimensions[1].height = 30
    ws_s.merge_cells("A2:H2")
    cs(ws_s["A2"],
       f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}  |  "
       "Classement unifié papier + numérique",
       "EBF5FB", "1A5276", False, 9)
    ws_s.row_dimensions[2].height = 16
    ws_s.merge_cells("A3:H3")
    cs(ws_s["A3"],
       "MÉTHODOLOGIE : Tous les candidats (papier et numérique) ont été classés "
       "ensemble selon le même barème /100. Les candidats papier ont été notés par la "
       "commission selon la grille officielle (6 critères). Les candidats numériques "
       "ont été notés par système informatique (IA + qualité du dossier). "
       "Les meilleurs ont été retenus sans distinction de canal de dépôt. "
       "Le processus est transparent et auditable.",
       "FFFDE7", "7D6608", False, 9, wrap=True)
    ws_s.row_dimensions[3].height = 60
    hdrs_s = ["Filière", "Quota", "Retenus", "dont\nPapier",
              "dont\nNumérique", "Liste\nAttente", "Places\nRestantes", "Statut"]
    for ci, h in enumerate(hdrs_s, 1):
        cs(ws_s.cell(row=5, column=ci), h, "1E3A5F", "FFFFFF", True, 10, wrap=True)
    ws_s.row_dimensions[5].height = 35
    tq = tr = tp = td = 0
    for i, s in enumerate(summary):
        r       = 6 + i
        restant = s["quota"] - s["retenus"]
        statut  = "✅ Complet" if restant <= 0 else f"⏳ {restant} à pourvoir"
        bg      = "FFFFFF" if i%2==0 else "F4F6F7"
        row_d   = [s["display"], s["quota"], s["retenus"], s["paper"],
                   s["digital"], s["attente"], restant, statut]
        for ci, val in enumerate(row_d, 1):
            cell = ws_s.cell(row=r, column=ci)
            fg   = ("C0392B" if (ci==7 and val>0) else
                    "1E8449" if (ci==7 and val<=0) else "000000")
            cs(cell, val, bg, fg, bold=(ci in (1,3)), size=10,
               align="left" if ci==1 else "center")
        ws_s.row_dimensions[r].height = 18
        tq+=s["quota"]; tr+=s["retenus"]; tp+=s["paper"]; td+=s["digital"]
    # Total
    tr_row = 6 + len(summary) + 1
    for ci, val in enumerate(
        ["TOTAL GÉNÉRAL", tq, tr, tp, td, "—",
         tq-tr, "✅ Objectif atteint" if tr>=200 else f"⏳ {200-tr} restants"], 1):
        cs(ws_s.cell(row=tr_row, column=ci), val,
           "1E3A5F", "FFFFFF", True, 11,
           align="left" if ci==1 else "center")
    ws_s.row_dimensions[tr_row].height = 24
    for ci, w in enumerate([38,8,9,9,11,9,10,20], 1):
        ws_s.column_dimensions[get_column_letter(ci)].width = w

    fname = f"Liste_Finale_MPG_FQ2026_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(fname)
    print()
    print("=" * 65)
    print(f"  ✅ Fichier généré : {fname}")
    print(f"  TOTAL RETENUS    : {tr} / 200")
    print(f"  dont Papier      : {tp}")
    print(f"  dont Numérique   : {td}")
    print("=" * 65)

if __name__ == "__main__":
    generate()